

from sklearn.model_selection import train_test_split
import os
import random
import imutils
import tensorflow as tf
import cv2
import numpy as np
import time
from skimage import morphology
import openpyxl
# Import thu vien segmentation_models
from segmentation_models.metrics import iou_score
from segmentation_models import Unet
import segmentation_models as sm
sm.set_framework("tf.keras")
sm.framework()
from matplotlib import pyplot as plt

# Dinh nghia bien
data_path  = "dataset"
w, h = 256, 256
batch_size = 16

# Dataset va Dataloader

BACKBONE = "resnet34"
preprocess_input = sm.get_preprocessing(BACKBONE)

# Dung de tao toan bo du lieu va load theo batch
class Dataset:
    def __init__(self, image_path, mask_path, w, h):
        # the paths of images
        self.image_path = image_path
        # the paths of segmentation images
        self.mask_path = mask_path

        self.w = w
        self.h = h

    def __getitem__(self, i):
        # read data
        image = cv2.imread(self.image_path[i])
        image = cv2.resize(image, (self.w, self.h), interpolation=cv2.INTER_AREA)
        image = preprocess_input(image)

        mask = cv2.imread(self.mask_path[i], cv2.IMREAD_UNCHANGED)
        image_mask = cv2.resize(mask, (self.w, self.h), interpolation=cv2.INTER_AREA)

        image_mask = [(image_mask == v) for v in [1]]
        image_mask = np.stack(image_mask, axis=-1).astype('float')

        return image, image_mask

class Dataloader(tf.keras.utils.Sequence):
    def __init__(self, dataset, batch_size,shape, shuffle=False):
        self.dataset = dataset
        self.batch_size = batch_size
        self.shuffle = shuffle
        self.shape = shape
        self.indexes = np.arange(self.shape)

    def __getitem__(self, i):
        # collect batch data
        start = i * self.batch_size
        stop = (i + 1) * self.batch_size
        data = []
        for j in range(start, stop):
            data.append(self.dataset[j])

        batch = [np.stack(samples, axis=0) for samples in zip(*data)]

        return tuple(batch)

    def __len__(self):
        return len(self.indexes) // self.batch_size

    def on_epoch_end(self):
        if self.shuffle:
            self.indexes = np.random.permutation(self.indexes)

#  Load thong tin tu folder dataset de tao 2 bien image_path, mask_path
def load_path(data_path):
    # Get normal image and mask
    classes = ['Class1']

    # Lop qua cac thu muc khong loi
    normal_image_path = []
    normal_mask_path = []
    for class_ in classes:
        current_folder = os.path.join(data_path, class_)
        for file in os.listdir(current_folder):
            if file.endswith("jpg") and (not file.startswith(".")):
                image_path = os.path.join(current_folder, file)
                mask_path = os.path.join(current_folder + "_mask", file)
                normal_mask_path.append(mask_path)
                normal_image_path.append(image_path)

    # Get defect image and mask
    defect_image_path = []
    defect_mask_path = []
    for class_ in classes:
        class_ = class_ + "_def"
        current_folder = os.path.join(data_path, class_)
        for file in os.listdir(current_folder):
            if file.endswith("jpg") and (not file.startswith(".")):
                image_path = os.path.join(current_folder, file)
                mask_path = os.path.join(current_folder + "_mask", file)
                defect_mask_path.append(mask_path)
                defect_image_path.append(image_path)

    # Normal:   normal_mask_path - chua toan bo duong dan cua mask
    #            normal_image_path-  chua toan bo duong dan den image
    # Defect:   defect_mask_path - chua toan bo duong dan cua mask
    #            defect_image_path-  chua toan bo duong dan den image

    idx = random.sample(range(len(normal_mask_path)), len(defect_mask_path))

    normal_mask_path_new = []
    normal_image_path_new = []

    for id in idx:
        normal_image_path_new.append(normal_image_path[id])
        normal_mask_path_new.append(normal_mask_path[id])

    image_path = normal_image_path_new + defect_image_path
    mask_path = normal_mask_path_new + defect_mask_path

    return image_path, mask_path

# Thu hien load va train model

# Load duong dan vao 2 bien
image_path, mask_path = load_path(data_path)

# Chia du lieu train, test
image_train, image_test, mask_train, mask_test = train_test_split(image_path, mask_path, test_size=0.01)

# Tao dataset va dataloader
train_dataset = Dataset(image_train, mask_train, w, h)
test_dataset = Dataset(image_test, mask_test, w, h)

train_loader = Dataloader(train_dataset, batch_size, shape=len(image_train), shuffle=True)
test_loader = Dataloader(test_dataset, batch_size, shape=len(image_test), shuffle=True)



# Khoi tao model
from tensorflow.keras.optimizers import Adam
from tensorflow.keras.callbacks import EarlyStopping
opt=tf.keras.optimizers.Adam(0.001)

#model= Unet(backbone_name='resnet34',classes=1,activation="sigmoid",input_shape=(256,256,3),encoder_freeze=True)
model= Unet(backbone_name='resnet34',encoder_weights="imagenet",classes=1,activation="sigmoid",input_shape=(256,256,3),encoder_freeze=True)
#model= Unet(input_shape=(256, 256, 3), classes=1, activation='sigmoid',  encoder_features='default', decoder_block_type='upsampling', decoder_filters=(256, 128, 64, 32, 16), decoder_use_batchnorm=True)
loss1 = sm.losses.categorical_focal_dice_loss
#model.compile(optimizer=opt,loss=loss1,metrics=[iou_score])
model.compile(optimizer="Adam",loss=loss1,metrics=[iou_score,'accuracy'])
# Train model
#is_train = False
is_train = True

if is_train:
    from keras.callbacks import ModelCheckpoint
    filepath="checkpoint.hdf5"
    callback = ModelCheckpoint(filepath, monitor='val_iou_score', verbose=1, save_best_only=True,mode='max')

    #history=model.fit_generator( train_loader, validation_data=test_loader, epochs=10, callbacks=[callback])
    history = model.fit_generator(train_loader, validation_data=test_loader,
                                  epochs=10, callbacks=[callback])
    final_accuracy = history.history["val_accuracy"][-5:]
    print("FINAL ACCURACY MEAN-5: ", np.mean(final_accuracy))
    print(history.history['accuracy'])
    print(history.history['iou_score'])
    plt.plot(history.history['accuracy'])
    plt.plot(history.history['iou_score'])
    plt.show()
else:
    # Load model de test
    row = 500
    column = 2
    j=0
    wb = openpyxl.Workbook()
    ws = wb.active
    model.load_weights("checkpoint.hdf5")
    test_path='TEST/Class1'
    test_mask='TEST/Class1_mask'
    test_folder = os.path.join(test_path)
    mask_folder = os.path.join(test_mask)
    for file in os.listdir(test_folder):
        if file.endswith("jpg"):
            print(file)
            j+=1
            # Read image file
            test_file = os.path.join(test_folder, file)
            mask_file = os.path.join(mask_folder, file)



            # Anh dau vao, ko phai mask
            image = cv2.imread(test_file)
            image1=image
            image1 = cv2.cvtColor(image1, cv2.COLOR_BGR2RGB)

            image = cv2.resize(image, (256, 256))
            # Dua qua model de predicted segmentation map
            img_expand = image[np.newaxis, ...]
            mask_predict = model.predict(image[np.newaxis, :, :, :])
            #prediction1 = model1.predict(image)
            # Doc mask thuc te
            image_mask = cv2.imread(mask_file, cv2.IMREAD_UNCHANGED)
            image_mask = cv2.resize(image_mask, (256, 256))



            plt.figure(figsize=(10, 10))
            plt.subplot(221)
            plt.title("Hình ảnh sản phẩm")
            plt.imshow(image1)
            plt.subplot(222)
            plt.title("Vết lỗi mong muốn")
            plt.imshow(image_mask)
            g = image_mask

            #g[g < 0.5] = 0
            #g[g >= 0.5] = 1







            start_time = time.time()


            z= mask_predict[0,:,:]
            z1 = mask_predict[0, :, :]


            z[z < 0.5] = 0
            z[z >= 0.5] = 1

            kernel = morphology.diamond(3)
            kernel2 = morphology.diamond(7)
            kernel3 = morphology.diamond(4)
            kernel4 = morphology.disk(8)
            kernel5 = morphology.disk(4)
            kernel6 = morphology.disk(3)

            #kernel2 = cv2.getGaussianKernel(10,5)
            h, w = image.shape[:2]
            mask = np.zeros((h + 2, w + 2), np.uint8)




            def fillhole(input_image):
                '''
                input gray binary image  get the filled image by floodfill method
                Note: only holes surrounded in the connected regions will be filled.
                :param input_image:
                :return:
                '''
                im_flood_fill = input_image.copy()
                h, w = input_image.shape[:2]
                mask = np.zeros((h + 2, w + 2), np.uint8)
                im_flood_fill = im_flood_fill.astype("uint8")
                cv2.floodFill(im_flood_fill, mask, (0, 0), 255)
                im_flood_fill_inv = cv2.bitwise_not(im_flood_fill)
                img_out = im_flood_fill_inv
                return img_out

            z = cv2.morphologyEx(z, cv2.MORPH_GRADIENT, kernel)
            z = cv2.dilate(z, kernel5)
            z = cv2.dilate(z, kernel5)
            z = fillhole(z)
            #z = cv2.dilate(z, kernel3)
            z = cv2.erode(z, kernel6)
            z = cv2.erode(z, kernel6)
            z = cv2.erode(z, kernel6)
            z = cv2.erode(z, kernel6)

            t=z
            I = (g & t)
            U = (g | t)
            I[I < 0.5] = 0
            I[I >= 0.5] = 1
            U[U < 0.5] = 0
            U[U >= 0.5] = 1
            nI = np.sum(I)
            nU = np.sum(U)
            IOU = nI / nU
            print(nI)
            print(nU)
            print(IOU)
            plt.subplot(223)
            plt.title("Vết lỗi dự đoán || IoU=" + str(IOU))
            plt.imshow(t)

            contours = cv2.findContours(z, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
            contours = imutils.grab_contours(contours)
            contours = sorted(contours, key=cv2.contourArea, reverse=False)

            number = 0
            # loop over our contours
            for c in contours:
                (x, y, w, h) = cv2.boundingRect(c)
                print(x, y, w, h)
                if (w > 6) or (h>6):
                # cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)
                # approximate the contour

                    cv2.rectangle(z, (x-5, y-5), (x + w+5, y + h+5), (255,0,0), 1)
                    number += 1

            print("Number of Contours found = " + str(number))
            if number >=1:
                note="Dự đoán: có lỗi"
            else:
                note="Dự đoán: không lỗi"

            end_time = time.time()
            elapsed_time = end_time - start_time


            plt.subplot(224)
            plt.title(note+"||Time= "+format(elapsed_time))
            v1=IOU
            v2=elapsed_time
            ws.cell(column=1, row=j + 1, value=v1)
            ws.cell(column=2, row=j + 1, value=v2)
            #plt.title(note + "||Time= " + round(elapsed_time,6))

            plt.imshow(z)
            #plt.show()
            plt.savefig(file)

    wb.save('TEST2.xlsx')
